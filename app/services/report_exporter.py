# 报表导出服务
import io
from datetime import datetime
from typing import List, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl.cell.cell


class ReportExporter:
    """报表导出基类"""
    
    def export(self, data: dict) -> bytes:
        raise NotImplementedError


class PDFExporter(ReportExporter):
    """PDF 报表导出"""
    
    def export(self, data: dict) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading2']
        normal_style = styles['Normal']
        
        elements = []
        
        elements.append(Paragraph("运行环境检查报表", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        
        if data.get("result_id"):
            elements.append(Paragraph(f"检查结果ID: {data['result_id']}", normal_style))
        if data.get("rule_name"):
            elements.append(Paragraph(f"检查规则: {data['rule_name']}", normal_style))
        if data.get("communication_name"):
            elements.append(Paragraph(f"通信机: {data['communication_name']}", normal_style))
        
        elements.append(Spacer(1, 0.5*cm))
        
        status = data.get("status", "unknown")
        status_text = {"success": "通过", "failed": "失败", "running": "进行中", "warning": "警告"}.get(status, status)
        elements.append(Paragraph(f"检查状态: {status_text}", normal_style))
        
        if data.get("start_time"):
            elements.append(Paragraph(f"开始时间: {data['start_time']}", normal_style))
        if data.get("end_time"):
            elements.append(Paragraph(f"结束时间: {data['end_time']}", normal_style))
        
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph("检查详情", heading_style))
        elements.append(Spacer(1, 0.3*cm))
        
        details = data.get("details", [])
        if details:
            table_data = [["检查项ID", "状态", "期望值", "实际值", "消息"]]
            
            for detail in details:
                status_map = {"pass": "通过", "fail": "失败", "error": "异常"}
                row = [
                    str(detail.get("check_item_id", "")),
                    status_map.get(detail.get("status", ""), detail.get("status", "")),
                    str(detail.get("expected_value", "-"))[:30],
                    str(detail.get("actual_value", "-"))[:30],
                    str(detail.get("message", "-"))[:40],
                ]
                table_data.append(row)
            
            table = Table(table_data, colWidths=[2*cm, 2*cm, 4*cm, 4*cm, 6*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("暂无检查详情", normal_style))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


class ExcelExporter(ReportExporter):
    """Excel 报表导出"""
    
    def export(self, data: dict) -> bytes:
        wb = Workbook()
        
        ws_summary = wb.active
        ws_summary.title = "检查结果汇总"
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws_summary.append(["运行环境检查报表"])
        ws_summary.merge_cells('A1:D1')
        
        summary_data = [
            ["生成时间:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["检查结果ID:", data.get("result_id", "-")],
            ["检查规则:", data.get("rule_name", "-")],
            ["通信机:", data.get("communication_name", "-")],
            ["检查状态:", data.get("status", "-")],
            ["开始时间:", data.get("start_time", "-")],
            ["结束时间:", data.get("end_time", "-")],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        ws_summary.append([])
        ws_summary.append(["检查详情"])
        
        headers = ["检查项ID", "状态", "期望值", "实际值", "消息"]
        ws_summary.append(headers)
        
        for cell in ws_summary[ws_summary.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        details = data.get("details", [])
        for detail in details:
            status_map = {"pass": "通过", "fail": "失败", "error": "异常"}
            row = [
                detail.get("check_item_id", ""),
                status_map.get(detail.get("status", ""), detail.get("status", "")),
                str(detail.get("expected_value", "-")),
                str(detail.get("actual_value", "-")),
                str(detail.get("message", "-")),
            ]
            ws_summary.append(row)
        
        for row in ws_summary.iter_rows(min_row=ws_summary.max_row - len(details), max_row=ws_summary.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border
        
        for col_idx, col in enumerate(ws_summary.columns, 1):
            max_length = 0
            from openpyxl.utils import get_column_letter
            column_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def export_check_result(result_id: int, format: str = "pdf") -> bytes:
    """导出检查结果"""
    from sqlalchemy import select
    from app.database import async_session_maker
    from app.models.check_result import CheckResult, CheckResultDetail
    from app.models import Communication
    from app.models.check_result import CheckRule
    
    import asyncio
    
    async def _export():
        async with async_session_maker() as db:
            result = await db.execute(select(CheckResult).where(CheckResult.id == result_id))
            check_result = result.scalar_one_or_none()
            
            if not check_result:
                raise ValueError("检查结果不存在")
            
            details_result = await db.execute(
                select(CheckResultDetail).where(CheckResultDetail.result_id == result_id)
            )
            details = details_result.scalars().all()
            
            rule_name = "-"
            if check_result.rule_id:
                rule_result = await db.execute(select(CheckRule).where(CheckRule.id == check_result.rule_id))
                rule = rule_result.scalar_one_or_none()
                if rule:
                    rule_name = rule.name
            
            comm_name = "-"
            if check_result.communication_id:
                comm_result = await db.execute(select(Communication).where(Communication.id == check_result.communication_id))
                comm = comm_result.scalar_one_or_none()
                if comm:
                    comm_name = comm.name
            
            data = {
                "result_id": result_id,
                "rule_name": rule_name,
                "communication_name": comm_name,
                "status": check_result.status,
                "start_time": check_result.start_time.strftime('%Y-%m-%d %H:%M:%S') if check_result.start_time else "-",
                "end_time": check_result.end_time.strftime('%Y-%m-%d %H:%M:%S') if check_result.end_time else "-",
                "details": [
                    {
                        "check_item_id": d.check_item_id,
                        "status": d.status,
                        "expected_value": d.expected_value,
                        "actual_value": d.actual_value,
                        "message": d.message,
                    }
                    for d in details
                ]
            }
            
            if format.lower() == "excel":
                return ExcelExporter().export(data)
            else:
                return PDFExporter().export(data)
    
    return asyncio.run(_export())
