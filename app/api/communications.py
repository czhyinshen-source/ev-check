# 通信机管理 API
from typing import Optional, List, Dict
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
import openpyxl

from app.database import get_db
from app.models import User, Communication, CommunicationGroup, SSHKey
from app.api.users import get_current_active_user
from app.utils.ssh_client import SSHClientWrapper
from app.services.rule_validator import ensure_not_in_execution_targets

router = APIRouter()


class CommunicationGroupCreate:
    """通信机分组创建"""
    name: str
    description: Optional[str] = None


class CommunicationGroupResponse:
    """通信机分组响应"""
    id: int
    name: str
    description: Optional[str]
    created_at: str
    updated_at: str


class CommunicationCreate:
    """通信机创建"""
    group_id: Optional[int] = None
    name: str
    ip_address: str
    port: int = 22
    username: str = "root"
    password: Optional[str] = None
    private_key_path: Optional[str] = None
    description: Optional[str] = None


class CommunicationResponse:
    """通信机响应"""
    id: int
    group_id: Optional[int]
    name: str
    ip_address: str
    port: int
    username: str
    description: Optional[str]
    is_active: bool
    created_at: str
    updated_at: str


class BatchUpdateRequest(BaseModel):
    """批量更新请求"""
    ids: List[int]
    group_id: Optional[int] = None
    description: Optional[str] = None


class BatchDeleteRequest(BaseModel):
    """批量删除请求"""
    ids: List[int]


class BatchTestRequest(BaseModel):
    """批量测试连接请求"""
    ids: List[int]


def _build_group_tree(groups: List[CommunicationGroup]) -> List[dict]:
    """将扁平分组列表构建为树形结构"""
    # 创建 ID 到分组的映射
    group_map = {}
    for g in groups:
        group_map[g.id] = {
            "id": g.id,
            "name": g.name,
            "parent_id": g.parent_id,
            "sort_order": g.sort_order,
            "description": g.description,
            "created_at": g.created_at.isoformat(),
            "updated_at": g.updated_at.isoformat(),
            "children": []
        }

    # 构建树形结构
    tree = []
    for group_id, group in group_map.items():
        if group["parent_id"] is None:
            tree.append(group)
        else:
            parent = group_map.get(group["parent_id"])
            if parent:
                parent["children"].append(group)

    # 对每层的子节点排序
    def sort_children(items):
        for item in items:
            if item["children"]:
                item["children"].sort(key=lambda x: x["sort_order"])
                sort_children(item["children"])

    sort_children(tree)
    tree.sort(key=lambda x: x["sort_order"])

    return tree


@router.get("/groups", response_model=List[dict])
async def list_groups(
    format: str = "flat",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """获取通信机分组列表

    format=flat: 返回扁平列表
    format=tree: 返回嵌套树形结构
    """
    result = await db.execute(select(CommunicationGroup).order_by(CommunicationGroup.sort_order))
    groups = result.scalars().all()

    if format == "tree":
        return _build_group_tree(list(groups))

    return [
        {
            "id": g.id,
            "name": g.name,
            "parent_id": g.parent_id,
            "sort_order": g.sort_order,
            "description": g.description,
            "created_at": g.created_at.isoformat(),
            "updated_at": g.updated_at.isoformat(),
        }
        for g in groups
    ]


@router.post("/groups", status_code=status.HTTP_201_CREATED)
async def create_group(
    group: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """创建通信机分组"""
    parent_id = group.get("parent_id")
    sort_order = group.get("sort_order", 0)

    # 检查同一父分组下名称是否重复
    if parent_id is not None:
        existing = await db.execute(
            select(CommunicationGroup).where(
                CommunicationGroup.name == group["name"],
                CommunicationGroup.parent_id == parent_id
            )
        )
    else:
        existing = await db.execute(
            select(CommunicationGroup).where(
                CommunicationGroup.name == group["name"],
                CommunicationGroup.parent_id.is_(None)
            )
        )

    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="同一父分组下已存在同名分组")

    db_group = CommunicationGroup(
        name=group["name"],
        parent_id=parent_id,
        sort_order=sort_order,
        description=group.get("description")
    )
    db.add(db_group)
    await db.commit()
    await db.refresh(db_group)
    return {
        "id": db_group.id,
        "name": db_group.name,
        "parent_id": db_group.parent_id,
        "sort_order": db_group.sort_order
    }


@router.put("/groups/{group_id}")
async def update_group(
    group_id: int,
    group_data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """更新通信机分组"""
    result = await db.execute(select(CommunicationGroup).where(CommunicationGroup.id == group_id))
    db_group = result.scalar_one_or_none()
    if not db_group:
        raise HTTPException(status_code=404, detail="分组不存在")

    # 检查循环引用：不能将分组设为自己的后代
    new_parent_id = group_data.get("parent_id")
    if new_parent_id is not None and new_parent_id == group_id:
        raise HTTPException(status_code=400, detail="不能将分组设为自己的父级")

    # 递归检查是否为自己的后代
    if new_parent_id:
        async def is_descendant(parent_id: int, child_id: int) -> bool:
            result = await db.execute(select(CommunicationGroup).where(CommunicationGroup.id == parent_id))
            parent = result.scalar_one_or_none()
            if not parent or parent.parent_id is None:
                return False
            if parent.parent_id == child_id:
                return True
            return await is_descendant(parent.parent_id, child_id)

        if await is_descendant(new_parent_id, group_id):
            raise HTTPException(status_code=400, detail="不能将分组移入自己的子分组")

    # 检查同一父分组下名称是否重复
    new_name = group_data.get("name")
    if new_name and new_name != db_group.name:
        if new_parent_id is not None:
            existing = await db.execute(
                select(CommunicationGroup).where(
                    CommunicationGroup.name == new_name,
                    CommunicationGroup.parent_id == new_parent_id,
                    CommunicationGroup.id != group_id
                )
            )
        else:
            existing = await db.execute(
                select(CommunicationGroup).where(
                    CommunicationGroup.name == new_name,
                    CommunicationGroup.parent_id.is_(None),
                    CommunicationGroup.id != group_id
                )
            )

        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="同一父分组下已存在同名分组")

    # 更新字段
    for key, value in group_data.items():
        if hasattr(db_group, key) and key not in ("id", "created_at"):
            setattr(db_group, key, value)

    await db.commit()
    await db.refresh(db_group)
    return {
        "id": db_group.id,
        "name": db_group.name,
        "parent_id": db_group.parent_id,
        "sort_order": db_group.sort_order
    }


@router.delete("/groups/{group_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_group(
    group_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """删除通信机分组"""
    from sqlalchemy.orm import selectinload
    result = await db.execute(
        select(CommunicationGroup)
        .options(selectinload(CommunicationGroup.children), selectinload(CommunicationGroup.communications))
        .where(CommunicationGroup.id == group_id)
    )
    db_group = result.scalar_one_or_none()
    if not db_group:
        raise HTTPException(status_code=404, detail="分组不存在")

    # 检查是否有子分组
    if db_group.children:
        raise HTTPException(status_code=400, detail="请先删除子分组")

    # 检查是否有通信机
    if db_group.communications:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="该分组下还有通信机，无法删除"
        )
        
    await ensure_not_in_execution_targets(db, "communication_group_id", group_id, f"节点分组 {db_group.name}")

    await db.delete(db_group)
    await db.commit()


@router.get("", response_model=List[dict])
async def list_communications(
    group_id: Optional[int] = None,
    page: int = 1,
    size: int = 20,
    q: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """获取通信机列表（支持分页和搜索）"""
    from sqlalchemy import func, or_
    from fastapi.responses import JSONResponse

    query = select(Communication)
    if group_id:
        # 获取所有分组以构建父子关系，实现级联查询
        all_groups_res = await db.execute(select(CommunicationGroup.id, CommunicationGroup.parent_id))
        all_groups = all_groups_res.all()
        
        from collections import defaultdict
        children_map = defaultdict(list)
        for g_id, p_id in all_groups:
            if p_id is not None:
                children_map[p_id].append(g_id)
                
        valid_group_ids = {group_id}
        stack = [group_id]
        while stack:
            curr = stack.pop()
            for child in children_map[curr]:
                if child not in valid_group_ids:
                    valid_group_ids.add(child)
                    stack.append(child)
                    
        query = query.where(Communication.group_id.in_(valid_group_ids))
        
    if q:
        query = query.where(
            or_(
                Communication.name.ilike(f"%{q}%"),
                Communication.ip_address.ilike(f"%{q}%"),
            )
        )

    # 计算总数
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar() or 0

    # 分页
    skip = (page - 1) * size
    query = query.order_by(Communication.id).offset(skip).limit(size)
    result = await db.execute(query)
    communications = result.scalars().all()

    data = [
        {
            "id": c.id,
            "group_id": c.group_id,
            "name": c.name,
            "ip_address": c.ip_address,
            "port": c.port,
            "username": c.username,
            "auth_method": c.auth_method,
            "description": c.description,
            "is_active": c.is_active,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        }
        for c in communications
    ]

    return JSONResponse(content=data, headers={"X-Total-Count": str(total)})



@router.post("", status_code=status.HTTP_201_CREATED)
async def create_communication(
    comm: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """创建通信机"""
    db_comm = Communication(
        group_id=comm.get("group_id"),
        name=comm["name"],
        ip_address=comm["ip_address"],
        port=comm.get("port", 22),
        username=comm.get("username", "root"),
        auth_method=comm.get("auth_method", "password"),
        password=comm.get("password"),
        private_key_path=comm.get("private_key_path"),
        description=comm.get("description"),
    )
    db.add(db_comm)
    await db.commit()
    await db.refresh(db_comm)
    return {
        "id": db_comm.id,
        "name": db_comm.name
    }


@router.put("/batch")
async def batch_update_communications(
    data: BatchUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """批量更新通信机"""
    if not data.ids:
        raise HTTPException(status_code=400, detail="未指定通信机")

    result = await db.execute(select(Communication).where(Communication.id.in_(data.ids)))
    communications = result.scalars().all()

    updated_count = 0
    for comm in communications:
        # 只更新明确提供的字段
        if data.group_id is not None:
            comm.group_id = data.group_id
            updated_count += 1
        if data.description is not None:
            comm.description = data.description
            updated_count += 1

    await db.commit()
    return {"updated": updated_count}


@router.delete("/batch", status_code=status.HTTP_204_NO_CONTENT)
async def batch_delete_communications(
    data: BatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """批量删除通信机"""
    if not data.ids:
        raise HTTPException(status_code=400, detail="未指定通信机")

    result = await db.execute(select(Communication).where(Communication.id.in_(data.ids)))
    communications = result.scalars().all()

    for com in communications:
        if not com:
            continue
        await ensure_not_in_execution_targets(db, "communication_id", com.id, f"通信机 {com.name}")
        await db.delete(com)

    await db.commit()


@router.post("/batch-test")
async def batch_test_connections(
    data: BatchTestRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """批量测试通信机连接"""
    if not data.ids:
        raise HTTPException(status_code=400, detail="未指定通信机")

    result = await db.execute(select(Communication).where(Communication.id.in_(data.ids)))
    communications = result.scalars().all()

    results = []
    online_count = 0
    offline_count = 0

    for comm in communications:
        status = "offline"
        message = "连接失败"

        try:
            private_key = None
            if comm.private_key_path and comm.private_key_path.startswith("key_"):
                ssh_key_id = int(comm.private_key_path.replace("key_", ""))
                ssh_key_result = await db.execute(select(SSHKey).where(SSHKey.id == ssh_key_id))
                ssh_key = ssh_key_result.scalar_one_or_none()
                if ssh_key:
                    private_key = ssh_key.private_key

            ssh_client = SSHClientWrapper(
                host=comm.ip_address,
                port=comm.port,
                username=comm.username,
                password=comm.password,
                private_key_path=None,
                private_key=private_key,
            )

            connected = await ssh_client.connect()
            await ssh_client.close()

            if connected:
                status = "online"
                message = "连接成功"
                online_count += 1
            else:
                offline_count += 1
        except Exception as e:
            message = f"连接失败: {str(e)}"
            offline_count += 1

        results.append({
            "id": comm.id,
            "name": comm.name,
            "ip_address": comm.ip_address,
            "status": status,
            "message": message
        })

    return {
        "results": results,
        "summary": {
            "total": len(results),
            "online": online_count,
            "offline": offline_count
        }
    }


@router.get("/excel-template")
async def download_excel_template(current_user: User = Depends(get_current_active_user)):
    """下载通信机导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "通信机导入模板"
    headers = ['名称', 'IP地址', '端口', '用户名', '描述', '通信机分组']
    ws.append(headers)
    
    ws.append(['示例机器', '192.168.1.100', 22, 'root', '这是一台示例机器', '本地'])
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename*=UTF-8''template.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


@router.get("/{comm_id}")
async def get_communication(
    comm_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """获取通信机详情"""
    result = await db.execute(select(Communication).where(Communication.id == comm_id))
    comm = result.scalar_one_or_none()
    if not comm:
        raise HTTPException(status_code=404, detail="通信机不存在")
    return {
        "id": comm.id,
        "group_id": comm.group_id,
        "name": comm.name,
        "ip_address": comm.ip_address,
        "port": comm.port,
        "username": comm.username,
        "auth_method": comm.auth_method,
        "password": comm.password,
        "private_key_path": comm.private_key_path,
        "description": comm.description,
        "is_active": comm.is_active,
    }


@router.put("/{comm_id}")
async def update_communication(
    comm_id: int,
    comm: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """更新通信机"""
    result = await db.execute(select(Communication).where(Communication.id == comm_id))
    db_comm = result.scalar_one_or_none()
    if not db_comm:
        raise HTTPException(status_code=404, detail="通信机不存在")

    for key, value in comm.items():
        if hasattr(db_comm, key):
            setattr(db_comm, key, value)

    await db.commit()
    await db.refresh(db_comm)
    return {"id": db_comm.id, "name": db_comm.name}


@router.get("/{comm_id}/status")
async def test_connection(
    comm_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """测试通信机连接"""
    result = await db.execute(select(Communication).where(Communication.id == comm_id))
    comm = result.scalar_one_or_none()
    if not comm:
        raise HTTPException(status_code=404, detail="通信机不存在")

    # 获取SSH密钥如果使用密钥认证）
    private_key = None
    if comm.private_key_path and comm.private_key_path.startswith("key_"):
        ssh_key_id = int(comm.private_key_path.replace("key_", ""))
        ssh_key_result = await db.execute(select(SSHKey).where(SSHKey.id == ssh_key_id))
        ssh_key = ssh_key_result.scalar_one_or_none()
        if ssh_key:
            private_key = ssh_key.private_key

    # 记录调试信息
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"测试通信机连接: id={comm_id}, host={comm.ip_address}, port={comm.port}, username={comm.username}")
    logger.info(f"认证方式: 密码={'是' if comm.password else '否'}, 密钥={'是' if private_key else '否'}")

    # 测试连接
    try:
        ssh_client = SSHClientWrapper(
            host=comm.ip_address,
            port=comm.port,
            username=comm.username,
            password=comm.password,
            private_key_path=None,
            private_key=private_key,
        )

        connected = await ssh_client.connect()
        await ssh_client.close()

        if connected:
            logger.info(f"通信机 {comm_id} 连接成功")
            return {"status": "online", "message": "连接成功"}
        else:
            logger.warning(f"通信机 {comm_id} 连接失败（connect 返回 False）")
            return {"status": "offline", "message": "连接失败：无法建立 SSH 连接"}
    except Exception as e:
        logger.error(f"通信机 {comm_id} 连接异常: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"status": "offline", "message": f"连接失败: {str(e)}"}


@router.delete("/{comm_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_communication(
    comm_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """删除通信机"""
    result = await db.execute(select(Communication).where(Communication.id == comm_id))
    comm = result.scalar_one_or_none()
    if not comm:
        raise HTTPException(status_code=404, detail="通信机不存在")

    await ensure_not_in_execution_targets(db, "communication_id", comm_id, f"通信机 {comm.name}")
    await db.delete(comm)
    await db.commit()



@router.post("/import-excel")
async def import_communications_from_excel(
    file: UploadFile = File(...),
    deploy_public_key: bool = False,
    ssh_key_id: Optional[int] = None,
    deploy_password: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """从Excel导入通信机"""
    print("收到Excel导入请求")
    print(f"文件名: {file.filename}")
    print(f"部署公钥: {deploy_public_key}")
    print(f"SSH密钥ID: {ssh_key_id}")
    print(f"部署密码: {deploy_password}")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 和 .xls 文件")

    # 读取Excel文件
    content = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(content))
    worksheet = workbook.active

    # 验证Excel格式 - 检查表头
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    expected_headers = ['名称', 'IP地址', '端口', '用户名', '描述', '通信机分组']
    
    # 验证表头是否匹配
    actual_headers = [str(h).strip() if h else '' for h in header_row[:6]]
    if actual_headers != expected_headers:
        raise HTTPException(
            status_code=400, 
            detail=f"Excel格式错误！表头应为：{', '.join(expected_headers)}"
        )

    # 获取所有通信机分组，创建名称到ID的映射
    group_result = await db.execute(select(CommunicationGroup))
    groups = group_result.scalars().all()
    group_name_to_id = {group.name: group.id for group in groups}
    
    # 解析Excel数据
    communications = []
    errors = []
    
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        # 验证必填字段
        name = row[0]
        ip_address = row[1]
        
        if not name:
            errors.append(f"第{row_num}行：通信机名称不能为空")
            continue
        if not ip_address:
            errors.append(f"第{row_num}行：IP地址不能为空")
            continue
        
        # 验证端口
        port = row[2]
        if port and not isinstance(port, int):
            try:
                port = int(port)
            except:
                errors.append(f"第{row_num}行：端口必须是数字")
                continue
        
        # 处理通信机分组
        group_name = row[5]
        group_id = None
        if group_name:
            group_name = str(group_name).strip()
            if group_name in group_name_to_id:
                group_id = group_name_to_id[group_name]
            else:
                errors.append(f"第{row_num}行：通信机分组 '{group_name}' 不存在")
                continue
        
        # 构建通信机数据
        comm_data = {
            "name": str(name).strip(),
            "ip_address": str(ip_address).strip(),
            "port": int(port) if port else 22,
            "username": str(row[3]).strip() if row[3] else "root",
            "description": str(row[4]).strip() if row[4] else None,
            "group_id": group_id,
        }
        communications.append(comm_data)

    if errors:
        raise HTTPException(status_code=400, detail="\n".join(errors))

    if not communications:
        raise HTTPException(status_code=400, detail="Excel文件中没有有效数据")

    # 批量创建通信机
    created_communications = []
    for comm_data in communications:
        db_comm = Communication(
            group_id=comm_data.get("group_id"),
            name=comm_data["name"],
            ip_address=comm_data["ip_address"],
            port=comm_data.get("port", 22),
            username=comm_data.get("username", "root"),
            description=comm_data.get("description"),
        )
        db.add(db_comm)
        await db.commit()
        await db.refresh(db_comm)
        created_communications.append(db_comm)

    # 如果需要部署公钥
    deployment_results = {"success": [], "failed": []}
    if deploy_public_key and ssh_key_id and deploy_password:
        # 获取SSH密钥
        result = await db.execute(select(SSHKey).where(SSHKey.id == ssh_key_id))
        ssh_key = result.scalar_one_or_none()
        if not ssh_key:
            raise HTTPException(status_code=404, detail="SSH密钥不存在")

        # 批量部署公钥
        for comm in created_communications:
            try:
                ssh_client = SSHClientWrapper(
                    host=comm.ip_address,
                    port=comm.port,
                    username=comm.username,
                    password=deploy_password,
                    private_key_path=None,
                    private_key=None,
                )

                connected = await ssh_client.connect()
                if connected:
                    command = f"mkdir -p ~/.ssh && chmod 700 ~/.ssh && echo '{ssh_key.public_key}' >> ~/.ssh/authorized_keys && chmod 600 ~/.ssh/authorized_keys"
                    exit_code, stdout, stderr = await ssh_client.execute(command)
                    
                    if exit_code == 0:
                        # 更新通信机使用SSH密钥
                        comm.private_key_path = f"key_{ssh_key_id}"
                        await db.commit()
                        deployment_results["success"].append(comm.name)
                    else:
                        deployment_results["failed"].append(f"{comm.name}: {stderr}")
                else:
                    deployment_results["failed"].append(f"{comm.name}: 无法连接")
            except Exception as e:
                deployment_results["failed"].append(f"{comm.name}: {str(e)}")
            finally:
                await ssh_client.close()

    return {
        "imported": len(created_communications),
        "deployment": deployment_results if deploy_public_key else None
    }


@router.post("/batch-deploy-ssh-key")
async def batch_deploy_ssh_key(
    deployment_data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """批量部署公钥到通信机"""
    communication_ids = deployment_data.get("communication_ids", [])
    ssh_key_id = deployment_data.get("ssh_key_id")
    password = deployment_data.get("password")

    if not communication_ids or not ssh_key_id or not password:
        raise HTTPException(status_code=400, detail="缺少必要参数")

    # 获取SSH密钥
    result = await db.execute(select(SSHKey).where(SSHKey.id == ssh_key_id))
    ssh_key = result.scalar_one_or_none()
    if not ssh_key:
        raise HTTPException(status_code=404, detail="SSH密钥不存在")

    # 获取通信机
    result = await db.execute(
        select(Communication).where(Communication.id.in_(communication_ids))
    )
    communications = result.scalars().all()

    if not communications:
        raise HTTPException(status_code=404, detail="没有找到指定的通信机")

    # 批量部署公钥
    deployment_results = {"success": [], "failed": []}
    for comm in communications:
        try:
            ssh_client = SSHClientWrapper(
                host=comm.ip_address,
                port=comm.port,
                username=comm.username,
                password=password,
                private_key_path=None,
                private_key=None,
            )

            connected = await ssh_client.connect()
            if connected:
                command = f"mkdir -p ~/.ssh && chmod 700 ~/.ssh && echo '{ssh_key.public_key}' >> ~/.ssh/authorized_keys && chmod 600 ~/.ssh/authorized_keys"
                exit_code, stdout, stderr = await ssh_client.execute(command)
                
                if exit_code == 0:
                    # 更新通信机使用SSH密钥
                    comm.private_key_path = f"key_{ssh_key_id}"
                    await db.commit()
                    deployment_results["success"].append(comm.name)
                else:
                    deployment_results["failed"].append(f"{comm.name}: {stderr}")
            else:
                deployment_results["failed"].append(f"{comm.name}: 无法连接")
        except Exception as e:
            deployment_results["failed"].append(f"{comm.name}: {str(e)}")
        finally:
            await ssh_client.close()

    return deployment_results
